# main.py — RareSignal AI FastAPI backend
import json
import uuid
import random
import math
from datetime import datetime, timedelta
from typing import List, Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session

from backend.database import init_db, get_db, Patient, SymptomEntry, BaselineProfile, ComputedSignal, FunctionalScore, Alert, SensorStream
from backend.schemas import (
    PatientCreate, PatientResponse,
    SymptomEntryCreate, SymptomEntryResponse,
    SignalRequest, SignalResponse,
    RiskPredictionRequest, RiskPredictionResponse,
    SummaryRequest, ClinicalSummaryResponse,
    DetailedReportRequest, DetailedReportResponse,
    HPOMatchRequest, HPOMatchResponse, HPOMatchResult,
    FlareAlertResponse,
    PatientHistoryResponse, SignalHistoryPoint,
    SensorStreamSummaryResponse,
    SharedExperienceSummaryResponse
)
from backend.signal_engine import (
    compute_all_signals, update_ewma_baseline, initialize_baseline_from_history
)
from backend.summary_generator import generate_structured_summary, generate_detailed_ai_report
from backend.disease_config import DISEASE_CONFIGS
from backend.hpo_matcher import match_symptom, get_disease_hpo_cluster


# ─── App Lifecycle ─────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    print("RareSignal AI backend started. DB initialized.")
    yield
    print("RareSignal AI backend shutting down.")


app = FastAPI(
    title="RareSignal AI",
    description="Structured Symptom-to-Signal Translation for Rare Disease Care",
    version="1.0.0",
    lifespan=lifespan
)

# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],  # restrict in production
#     allow_credentials=True,
#     allow_methods=["*"],
#     allow_headers=["*"]
# )

# ─── CORS Configuration ─────────────────────────────────────────────

origins = [
    "http://localhost:3000",                         # React dev server
    "https://hackrare-project.vercel.app",           # Vercel production
    "https://hackrare-project-*.vercel.app",         # Vercel preview deployments
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Utility ───────────────────────────────────────────────────────────────────

def _get_patient_or_404(patient_id: str, db: Session) -> Patient:
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail=f"Patient '{patient_id}' not found")
    return patient


def _entries_to_dicts(entries: List[SymptomEntry]) -> List[dict]:
    return [{
        "timestamp": e.timestamp,
        "symptoms": e.symptoms,
        "triggers": e.triggers,
        "lifestyle_context": e.lifestyle_context,
        "shared_experience": e.shared_experience,
        "notes": e.notes
    } for e in entries]


def _get_or_init_baseline(patient_id: str, db: Session) -> BaselineProfile:
    baseline = db.query(BaselineProfile).filter(BaselineProfile.patient_id == patient_id).first()
    if not baseline:
        baseline = BaselineProfile(
            patient_id=patient_id,
            mu_json="{}",
            sigma_json="{}",
            n_observations=0
        )
        db.add(baseline)
        db.commit()
        db.refresh(baseline)
    return baseline


def _update_baseline_from_entry(baseline: BaselineProfile, symptoms: dict, db: Session):
    """Update EWMA baseline with new symptom entry."""
    mu = baseline.mu
    sigma = baseline.sigma
    alpha = baseline.alpha

    for sym, val in symptoms.items():
        mu_prev = mu.get(sym, val)
        sig_prev = sigma.get(sym, 2.0)
        mu_new, sig_new = update_ewma_baseline(mu_prev, sig_prev, val, alpha)
        mu[sym] = round(mu_new, 4)
        sigma[sym] = round(sig_new, 4)

    baseline.mu_json = json.dumps(mu)
    baseline.sigma_json = json.dumps(sigma)
    baseline.n_observations += 1
    baseline.last_updated = datetime.utcnow()
    db.commit()


AUTONOMIC_DISEASES = {"POTS", "EDS", "ENS", "FXS", "HD", "PKU", "NF1", "PRION"}
RESPIRATORY_DISEASES = {"PCD", "CF", "Heterotaxy", "RRP", "SMA", "Pompe", "TS", "RTT"}
INFLAMMATORY_DISEASES = {"FMF", "Gaucher", "WD", "Alkaptonuria"}
CARDIOVASCULAR_DISEASES = {"MFS", "Heterotaxy", "POTS"}


def _relevant_lifestyle_keys_for_disease(disease: str) -> List[str]:
    if disease in AUTONOMIC_DISEASES:
        return [
            "sleep_duration_hours", "sleep_disruption", "mentally_demanding_day",
            "overexertion", "activity_worsened_symptoms", "hydration_level",
            "heat_exposure"
        ]
    if disease in RESPIRATORY_DISEASES:
        return [
            "sleep_duration_hours", "sleep_disruption", "activity_level",
            "overexertion", "activity_worsened_symptoms", "cold_exposure",
            "heat_exposure", "illness_symptoms", "travel_or_routine_change"
        ]
    if disease in INFLAMMATORY_DISEASES:
        return [
            "sleep_duration_hours", "sleep_disruption", "mentally_demanding_day",
            "hydration_level", "large_or_unusual_meals", "missed_meals",
            "illness_symptoms", "cold_exposure"
        ]
    if disease in CARDIOVASCULAR_DISEASES:
        return [
            "sleep_duration_hours", "sleep_disruption", "activity_level",
            "overexertion", "activity_worsened_symptoms", "hydration_level",
            "heat_exposure"
        ]
    return [
        "sleep_duration_hours", "sleep_disruption", "activity_level",
        "mentally_demanding_day", "hydration_level", "illness_symptoms"
    ]


def _label_for_feature(feature_key: str, config: dict) -> str:
    symptom_labels = config.get("symptom_labels", {})
    lifestyle_labels = {
        "sleep_duration_hours": "Sleep duration",
        "sleep_disruption": "Sleep disruption",
        "activity_level": "Activity level",
        "overexertion": "Overexertion",
        "activity_worsened_symptoms": "Activity worsened symptoms",
        "mentally_demanding_day": "Mentally demanding day",
        "hydration_level": "Hydration level",
        "large_or_unusual_meals": "Large or unusual meals",
        "missed_meals": "Missed meals",
        "heat_exposure": "Heat exposure",
        "cold_exposure": "Cold exposure",
        "illness_symptoms": "Illness symptoms",
        "travel_or_routine_change": "Travel or routine change",
        "emotional_strain_note": "Emotional strain",
    }
    if feature_key in symptom_labels:
        return symptom_labels[feature_key]
    if feature_key in lifestyle_labels:
        return lifestyle_labels[feature_key]
    return feature_key.replace("_", " ").title()


def _lifestyle_factor_score(key: str, value) -> float:
    if value is None:
        return 0.0
    if key == "sleep_duration_hours":
        hours = float(value)
        if hours < 5:
            return 1.0
        if hours < 6.5:
            return 0.65
        if hours > 10:
            return 0.3
        return 0.0
    if isinstance(value, str):
        if value == "yes":
            return 1.0
        if value == "unsure":
            return 0.35
        if value == "high":
            return 0.8
        if value == "moderate":
            return 0.35
        if value == "low":
            return 0.6 if key == "hydration_level" else 0.0
        if value == "adequate":
            return 0.0
        if value == "no":
            return 0.0
    return 0.0


def _softmax_probability_map(scores: dict) -> dict:
    exp_values = {k: math.exp(v) for k, v in scores.items()}
    total = sum(exp_values.values()) or 1.0
    return {k: round(v / total, 4) for k, v in exp_values.items()}


def _compute_disease_aware_risk(patient: Patient, latest_entry: SymptomEntry, latest_signal: ComputedSignal, latest_fis: Optional[FunctionalScore]) -> dict:
    config = DISEASE_CONFIGS.get(patient.disease, {})
    all_symptoms = config.get("symptoms", [])
    primary_symptoms = [s for s in all_symptoms if s not in {"sleep_quality", "stress_symptom_severity"}][:5]
    relevant_symptoms = primary_symptoms + [s for s in ["sleep_quality", "stress_symptom_severity"] if s in all_symptoms]
    relevant_triggers = list(config.get("triggers", []))
    relevant_lifestyle_keys = _relevant_lifestyle_keys_for_disease(patient.disease)

    z_scores = latest_signal.z_scores or {}
    symptom_contribs = []
    for sym in relevant_symptoms:
        z_val = z_scores.get(sym)
        if z_val is None:
            continue
        contribution = max(0.0, min(abs(float(z_val)) / 3.0, 1.2))
        if contribution > 0:
            symptom_contribs.append({
                "feature": _label_for_feature(sym, config),
                "importance": round(min(contribution, 1.0), 3),
                "direction": "+" if float(z_val) >= 0 else "-"
            })

    active_triggers = [t for t in latest_entry.triggers if t in relevant_triggers]
    trigger_contribs = [{
        "feature": _label_for_feature(trigger, config),
        "importance": 0.18,
        "direction": "+"
    } for trigger in active_triggers]

    lifestyle_context = latest_entry.lifestyle_context or {}
    lifestyle_contribs = []
    for key in relevant_lifestyle_keys:
        score = _lifestyle_factor_score(key, lifestyle_context.get(key))
        if score <= 0:
            continue
        lifestyle_contribs.append({
            "feature": _label_for_feature(key, config),
            "importance": round(min(score * 0.35, 1.0), 3),
            "direction": "+"
        })

    symptom_pressure = min(1.0, sum(item["importance"] for item in symptom_contribs) / max(len(relevant_symptoms), 1) * 1.4)
    trigger_pressure = min(1.0, len(active_triggers) / max(3, len(relevant_triggers) or 1))
    lifestyle_pressure = min(1.0, sum(item["importance"] for item in lifestyle_contribs))
    volatility_pressure = min(1.0, (latest_signal.volatility_index or 0.0) / 0.9)
    fis_pressure = min(1.0, (latest_fis.fis_composite if latest_fis else 0.0) / 0.85)
    red_flag_pressure = 1.0 if latest_signal.risk_category == "CRITICAL" else 0.0

    composite_pressure = (
        symptom_pressure * 0.42 +
        trigger_pressure * 0.14 +
        lifestyle_pressure * 0.18 +
        volatility_pressure * 0.14 +
        fis_pressure * 0.12
    )
    composite_pressure = min(1.3, composite_pressure + red_flag_pressure * 0.35)

    scores = {
        "LOW": 2.4 - (composite_pressure * 3.2),
        "MODERATE": 0.9 + (composite_pressure * 1.6) - abs(composite_pressure - 0.45),
        "HIGH": (composite_pressure * 2.4) - 0.5,
        "CRITICAL": (composite_pressure * 3.0) - 1.8 + red_flag_pressure
    }
    probs = _softmax_probability_map(scores)
    risk_category = max(probs.items(), key=lambda item: item[1])[0]

    ranked_features = sorted(
        symptom_contribs + trigger_contribs + lifestyle_contribs,
        key=lambda item: item["importance"],
        reverse=True
    )[:5]
    if not ranked_features:
        ranked_features = [{
            "feature": "Recent disease-specific inputs",
            "importance": round(composite_pressure, 3),
            "direction": "+"
        }]

    context_labels = [item["feature"] for item in lifestyle_contribs[:3] + trigger_contribs[:2]]
    if context_labels:
        context_statement = (
            "Relevant recent context for this disease included "
            + ", ".join(context_labels)
            + ". These factors are shown as associative context and not as causes."
        )
    else:
        context_statement = "No major disease-relevant lifestyle or trigger factors were prominent in the latest check-in."

    return {
        "risk_category": risk_category,
        "risk_probabilities": probs,
        "top_contributing_features": ranked_features,
        "context_statement": context_statement,
        "composite_pressure": composite_pressure,
    }


def _sanitize_shared_snippet(text: str) -> Optional[str]:
    if not text:
        return None
    cleaned = " ".join(str(text).strip().split())
    if len(cleaned) < 8:
        return None
    return cleaned[:160]


def _sensor_modifiers_for_disease(disease: str) -> dict:
    return {
        "POTS": {"hr_shift": 9, "hrv_shift": -10, "spo2_shift": -0.2, "stress_shift": 10},
        "PCD": {"hr_shift": 4, "hrv_shift": -6, "spo2_shift": -1.2, "stress_shift": 8},
        "Heterotaxy": {"hr_shift": 5, "hrv_shift": -8, "spo2_shift": -1.8, "stress_shift": 9},
        "ENS": {"hr_shift": 2, "hrv_shift": -3, "spo2_shift": 0.0, "stress_shift": 7},
        "EDS": {"hr_shift": 4, "hrv_shift": -5, "spo2_shift": -0.1, "stress_shift": 6},
        "FMF": {"hr_shift": 6, "hrv_shift": -7, "spo2_shift": -0.1, "stress_shift": 9},
        "CF": {"hr_shift": 6, "hrv_shift": -6, "spo2_shift": -1.5, "stress_shift": 8},
        "HD": {"hr_shift": 3, "hrv_shift": -8, "spo2_shift": -0.1, "stress_shift": 9},
        "RTT": {"hr_shift": 5, "hrv_shift": -7, "spo2_shift": -1.0, "stress_shift": 9},
        "MFS": {"hr_shift": 6, "hrv_shift": -5, "spo2_shift": -0.3, "stress_shift": 7},
        "SMA": {"hr_shift": 4, "hrv_shift": -6, "spo2_shift": -1.1, "stress_shift": 7},
        "FXS": {"hr_shift": 3, "hrv_shift": -9, "spo2_shift": 0.0, "stress_shift": 11},
        "NF1": {"hr_shift": 3, "hrv_shift": -5, "spo2_shift": -0.1, "stress_shift": 8},
        "PKU": {"hr_shift": 2, "hrv_shift": -6, "spo2_shift": 0.0, "stress_shift": 8},
        "WD": {"hr_shift": 5, "hrv_shift": -7, "spo2_shift": -0.2, "stress_shift": 8},
        "Pompe": {"hr_shift": 5, "hrv_shift": -6, "spo2_shift": -0.9, "stress_shift": 7},
        "TS": {"hr_shift": 4, "hrv_shift": -7, "spo2_shift": -1.1, "stress_shift": 8},
        "Gaucher": {"hr_shift": 4, "hrv_shift": -5, "spo2_shift": -0.2, "stress_shift": 7},
        "Alkaptonuria": {"hr_shift": 3, "hrv_shift": -5, "spo2_shift": -0.1, "stress_shift": 7},
        "Achondroplasia": {"hr_shift": 4, "hrv_shift": -4, "spo2_shift": -0.8, "stress_shift": 6},
        "RRP": {"hr_shift": 5, "hrv_shift": -5, "spo2_shift": -1.0, "stress_shift": 8},
        "PRION": {"hr_shift": 6, "hrv_shift": -10, "spo2_shift": -0.6, "stress_shift": 11},
    }.get(disease, {"hr_shift": 3, "hrv_shift": -4, "spo2_shift": -0.2, "stress_shift": 5})


def _build_sensor_insights(disease: str, metrics: dict) -> tuple[list, list]:
    insights = []
    alerts = []

    autonomic_diseases = {"POTS", "EDS", "ENS", "FXS", "HD", "PKU", "NF1", "PRION"}
    respiratory_diseases = {"PCD", "CF", "Heterotaxy", "RRP", "SMA", "Pompe", "TS", "RTT"}
    inflammatory_diseases = {"FMF", "Gaucher", "WD", "Alkaptonuria"}
    cardiovascular_diseases = {"MFS", "Heterotaxy", "POTS"}

    if disease in autonomic_diseases:
        if metrics["hrv_rmssd"] <= 24:
            insights.append("Autonomic recovery markers were suppressed, with lower HRV suggesting higher nervous-system strain.")
            alerts.append("Autonomic stress pattern detected")
        else:
            insights.append("Autonomic variability stayed in a moderate range, without marked suppression in HRV.")
        if metrics["stress_load"] >= 70:
            insights.append("Wearable stress load remained elevated, which may align with symptom-triggering autonomic burden.")
            alerts.append("High autonomic stress load")

    if disease in respiratory_diseases:
        if metrics["spo2_avg"] is not None and metrics["spo2_avg"] < 95:
            insights.append("Oxygen saturation trended below the preferred range, which is relevant for respiratory burden tracking.")
            alerts.append("Respiratory oxygen variability detected")
        else:
            insights.append("Oxygen saturation remained relatively stable through the replayed respiratory monitoring window.")
        if metrics["heart_rate_avg"] >= 90:
            insights.append("Average heart rate rose alongside the respiratory profile, which can indicate higher breathing effort.")

    if disease in inflammatory_diseases:
        if metrics["skin_temp_avg"] >= 37.0:
            insights.append("Skin temperature trended above the expected wearable baseline, which can correlate with inflammatory or flare activity.")
            alerts.append("Temperature elevation detected")
        else:
            insights.append("Skin temperature stayed close to the expected wearable baseline during this replay window.")
        if metrics["stress_load"] >= 68:
            insights.append("Stress load and recovery imbalance may reflect physiologic strain during an inflammatory-heavy day.")

    if disease in cardiovascular_diseases:
        if metrics["heart_rate_avg"] >= 92:
            insights.append("Average heart rate remained elevated over the replay window, suggesting higher cardiovascular strain.")
            alerts.append("Elevated cardiovascular load")
        else:
            insights.append("Average heart rate stayed near the expected wearable baseline for this cardiovascular profile.")
        if metrics["heart_rate_max"] >= 120:
            insights.append("Peak heart-rate excursions were notable and may be relevant to symptom flare review.")

    if disease not in autonomic_diseases | respiratory_diseases | inflammatory_diseases | cardiovascular_diseases:
        if metrics["activity_load"] >= 75 and metrics["recovery_score"] <= 45:
            insights.append("Activity burden exceeded recovery capacity, which may align with fatigue-heavy or mobility-limited days.")
            alerts.append("Recovery mismatch after activity")
        else:
            insights.append("Activity and recovery stayed in a moderate range during the replayed monitoring window.")
        if metrics["heart_rate_avg"] >= 90:
            insights.append("Average heart rate was mildly elevated compared with the expected wearable baseline.")

    if not alerts:
        alerts.append("No critical wearable-derived alerts detected")

    return insights, alerts


def _create_sensor_stream_record(patient: Patient, db: Session, existing_stream: Optional[SensorStream] = None) -> SensorStream:
    seed = f"{patient.id}:{patient.disease}:{patient.wearable_device_type or 'generic'}"
    rng = random.Random(seed)
    mods = _sensor_modifiers_for_disease(patient.disease)

    heart_rate_avg = round(78 + mods["hr_shift"] + rng.uniform(-4, 6), 1)
    heart_rate_resting = round(max(52, heart_rate_avg - rng.uniform(10, 18)), 1)
    heart_rate_max = round(heart_rate_avg + rng.uniform(22, 48), 1)
    hrv_rmssd = round(max(10, 34 + mods["hrv_shift"] + rng.uniform(-7, 7)), 1)
    spo2_avg = round(min(99.0, max(90.0, 97.2 + mods["spo2_shift"] + rng.uniform(-1.2, 0.6))), 1)
    skin_temp_avg = round(36.6 + rng.uniform(-0.4, 0.8), 1)
    activity_load = round(min(100.0, max(10.0, 52 + rng.uniform(-18, 28))), 1)
    recovery_score = round(min(100.0, max(5.0, 58 + rng.uniform(-20, 18) - (mods["stress_shift"] * 0.6))), 1)
    stress_load = round(min(100.0, max(5.0, 44 + mods["stress_shift"] + rng.uniform(-10, 18))), 1)

    signal_quality = "GOOD"
    if hrv_rmssd <= 18 or spo2_avg <= 93:
        signal_quality = "REVIEW"
    elif stress_load >= 78:
        signal_quality = "WATCH"

    metrics = {
        "heart_rate_avg": heart_rate_avg,
        "heart_rate_resting": heart_rate_resting,
        "heart_rate_max": heart_rate_max,
        "hrv_rmssd": hrv_rmssd,
        "spo2_avg": spo2_avg,
        "skin_temp_avg": skin_temp_avg,
        "activity_load": activity_load,
        "recovery_score": recovery_score,
        "stress_load": stress_load,
    }
    insights, alerts = _build_sensor_insights(patient.disease, metrics)

    stream = existing_stream or SensorStream(patient_id=patient.id)
    stream.source = "simulated_replay"
    stream.recorded_at = datetime.utcnow()
    stream.sample_count = 1440
    stream.duration_minutes = 24 * 60
    stream.heart_rate_avg = heart_rate_avg
    stream.heart_rate_resting = heart_rate_resting
    stream.heart_rate_max = heart_rate_max
    stream.hrv_rmssd = hrv_rmssd
    stream.spo2_avg = spo2_avg
    stream.skin_temp_avg = skin_temp_avg
    stream.activity_load = activity_load
    stream.recovery_score = recovery_score
    stream.stress_load = stress_load
    stream.signal_quality = signal_quality
    stream.insights_json = json.dumps(insights)
    stream.alerts_json = json.dumps(alerts)
    if existing_stream is None:
        db.add(stream)
    db.commit()
    db.refresh(stream)
    return stream


# ─── Patients ──────────────────────────────────────────────────────────────────

@app.post("/patients", response_model=PatientResponse, tags=["Patients"])
def create_patient(body: PatientCreate, db: Session = Depends(get_db)):
    if body.disease not in DISEASE_CONFIGS:
        raise HTTPException(400, f"Unknown disease: {body.disease}")
    existing = db.query(Patient).filter(Patient.id == body.id).first()
    if existing:
        raise HTTPException(409, f"Patient '{body.id}' already exists")
    patient = Patient(
        id=body.id,
        disease=body.disease,
        uses_wearable=body.uses_wearable,
        wearable_device_type=body.wearable_device_type,
        wants_wearable_link=body.wants_wearable_link
    )
    db.add(patient)
    # Initialize baseline profile
    baseline = BaselineProfile(patient_id=body.id, mu_json="{}", sigma_json="{}")
    db.add(baseline)
    db.commit()
    db.refresh(patient)
    if patient.uses_wearable and patient.wants_wearable_link:
        _create_sensor_stream_record(patient, db)
    return patient


@app.get("/patients/{patient_id}", response_model=PatientResponse, tags=["Patients"])
def get_patient(patient_id: str, db: Session = Depends(get_db)):
    return _get_patient_or_404(patient_id, db)


@app.get("/patients", response_model=List[PatientResponse], tags=["Patients"])
def list_patients(db: Session = Depends(get_db)):
    return db.query(Patient).all()


@app.get("/patients/{patient_id}/sensor-summary", response_model=SensorStreamSummaryResponse, tags=["Sensors"])
def get_sensor_summary(patient_id: str, db: Session = Depends(get_db)):
    patient = _get_patient_or_404(patient_id, db)
    if not patient.uses_wearable or not patient.wants_wearable_link:
        return SensorStreamSummaryResponse(
            patient_id=patient_id,
            linked=False,
            device_type=patient.wearable_device_type,
            insights=[],
            alerts=[]
        )

    stream = (
        db.query(SensorStream)
        .filter(SensorStream.patient_id == patient_id)
        .order_by(SensorStream.recorded_at.desc())
        .first()
    )
    if not stream:
        stream = _create_sensor_stream_record(patient, db)
    else:
        stream = _create_sensor_stream_record(patient, db, existing_stream=stream)

    return SensorStreamSummaryResponse(
        patient_id=patient_id,
        linked=True,
        device_type=patient.wearable_device_type,
        stream_source=stream.source,
        collected_at=stream.recorded_at,
        sample_count=stream.sample_count,
        duration_minutes=stream.duration_minutes,
        heart_rate_avg=stream.heart_rate_avg,
        heart_rate_resting=stream.heart_rate_resting,
        heart_rate_max=stream.heart_rate_max,
        hrv_rmssd=stream.hrv_rmssd,
        spo2_avg=stream.spo2_avg,
        skin_temp_avg=stream.skin_temp_avg,
        activity_load=stream.activity_load,
        recovery_score=stream.recovery_score,
        stress_load=stream.stress_load,
        signal_quality=stream.signal_quality,
        insights=stream.insights,
        alerts=stream.alerts
    )


# ─── Symptom Entries ───────────────────────────────────────────────────────────

@app.post("/entries", response_model=SymptomEntryResponse, tags=["Entries"])
def add_symptom_entry(body: SymptomEntryCreate, db: Session = Depends(get_db)):
    patient = _get_patient_or_404(body.patient_id, db)
    config = DISEASE_CONFIGS[patient.disease]
    required_global_symptoms = ["sleep_quality", "stress_symptom_severity"]

    # Validate symptom names
    valid_symptoms = config.get("symptoms", [])
    for sym in body.symptoms:
        if sym not in valid_symptoms:
            raise HTTPException(400, f"Unknown symptom '{sym}' for disease {patient.disease}")

    missing_required = [sym for sym in required_global_symptoms if sym in valid_symptoms and sym not in body.symptoms]
    if missing_required:
        raise HTTPException(
            400,
            f"Missing required symptom values: {', '.join(missing_required)}"
        )

    # Validate symptom range
    for sym, val in body.symptoms.items():
        if not (0.0 <= val <= 10.0):
            raise HTTPException(400, f"Symptom '{sym}' value {val} out of range [0, 10]")

    timestamp = body.timestamp or datetime.utcnow()
    entry = SymptomEntry(
        patient_id=body.patient_id,
        timestamp=timestamp,
        symptoms_json=json.dumps(body.symptoms),
        triggers_json=json.dumps(body.triggers),
        lifestyle_json=json.dumps(body.lifestyle_context or {}),
        shared_experience_json=json.dumps(body.shared_experience or {}),
        notes=body.notes
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)

    # Update baseline
    baseline = _get_or_init_baseline(body.patient_id, db)
    _update_baseline_from_entry(baseline, body.symptoms, db)

    return SymptomEntryResponse(
        id=entry.id,
        patient_id=entry.patient_id,
        timestamp=entry.timestamp,
        symptoms=entry.symptoms,
        triggers=entry.triggers,
        lifestyle_context=entry.lifestyle_context,
        shared_experience=entry.shared_experience,
        notes=entry.notes
    )


# ─── CSV / Spreadsheet Upload ──────────────────────────────────────────────────

@app.post("/upload-symptom-csv/{patient_id}", tags=["Entries"])
async def upload_symptom_csv(
    patient_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Bulk-import symptom entries from a CSV or Excel file.
    Expected columns (case-insensitive):
      date | symptom_<name> (multiple) | triggers | notes
    OR spreadsheet format:
      date | symptom_name | value | triggers | notes  (one row per symptom)
    Returns { imported, skipped, errors }
    """
    import io, csv
    patient = _get_patient_or_404(patient_id, db)
    config = DISEASE_CONFIGS.get(patient.disease, {})
    valid_symptoms = set(config.get("symptoms", []))

    content = await file.read()
    filename = (file.filename or "").lower()

    # ── Parse CSV or Excel ──
    rows = []
    if filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [v for v in row])))
        except ImportError:
            raise HTTPException(400, "openpyxl not installed. Use CSV or run: pip install openpyxl")
    else:
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): v for k, v in row.items()})

    imported = 0
    errors = []

    # Detect format
    headers_set = set(rows[0].keys()) if rows else set()
    is_long_format = "symptom_name" in headers_set and "value" in headers_set

    if is_long_format:
        # Long format: one row per (date, symptom, value)
        by_date: dict = {}
        for i, row in enumerate(rows, start=2):
            try:
                date_str = str(row.get("date", "")).strip()
                sym = str(row.get("symptom_name", "")).strip().lower().replace(" ", "_")
                val = float(row.get("value", 0))
                triggers_raw = str(row.get("triggers", "")).strip()
                notes = str(row.get("notes", "")).strip() or None
                if sym not in valid_symptoms:
                    errors.append(f"Row {i}: Unknown symptom '{sym}'")
                    continue
                if not (0 <= val <= 10):
                    errors.append(f"Row {i}: Value {val} out of range for '{sym}'")
                    continue
                key = date_str or datetime.utcnow().isoformat()
                if key not in by_date:
                    by_date[key] = {"symptoms": {}, "triggers": [], "notes": notes}
                by_date[key]["symptoms"][sym] = round(val, 1)
                if triggers_raw:
                    by_date[key]["triggers"] = [t.strip() for t in triggers_raw.split(",") if t.strip()]
            except Exception as e:
                errors.append(f"Row {i}: {e}")

        for date_str, data in by_date.items():
            try:
                ts = datetime.fromisoformat(date_str) if date_str else datetime.utcnow()
            except ValueError:
                ts = datetime.utcnow()
            entry = SymptomEntry(
                patient_id=patient_id,
                timestamp=ts,
                symptoms_json=json.dumps(data["symptoms"]),
                triggers_json=json.dumps(data["triggers"]),
                lifestyle_json=json.dumps({}),
                shared_experience_json=json.dumps({}),
                notes=data["notes"],
            )
            db.add(entry); db.commit(); db.refresh(entry)
            baseline = _get_or_init_baseline(patient_id, db)
            _update_baseline_from_entry(baseline, data["symptoms"], db)
            imported += 1
    else:
        # Wide format: columns are symptom_<name>
        sym_cols = {k: k.replace("symptom_", "") for k in headers_set if k.startswith("symptom_")}
        for i, row in enumerate(rows, start=2):
            try:
                date_str = str(row.get("date", "")).strip()
                ts = datetime.fromisoformat(date_str) if date_str else datetime.utcnow()
            except ValueError:
                ts = datetime.utcnow()
            symptoms: dict = {}
            row_errors = []
            for col, sym in sym_cols.items():
                if sym not in valid_symptoms:
                    continue
                raw = row.get(col, "")
                if raw is None or str(raw).strip() == "":
                    continue
                try:
                    val = float(raw)
                    if 0 <= val <= 10:
                        symptoms[sym] = round(val, 1)
                    else:
                        row_errors.append(f"Row {i}: {sym}={val} out of range")
                except (ValueError, TypeError):
                    row_errors.append(f"Row {i}: Cannot parse {col}='{raw}'")
            errors.extend(row_errors)
            if not symptoms:
                continue
            triggers_raw = str(row.get("triggers", "")).strip()
            triggers = [t.strip() for t in triggers_raw.split(",") if t.strip()] if triggers_raw else []
            notes = str(row.get("notes", "")).strip() or None
            entry = SymptomEntry(
                patient_id=patient_id,
                timestamp=ts,
                symptoms_json=json.dumps(symptoms),
                triggers_json=json.dumps(triggers),
                lifestyle_json=json.dumps({}),
                shared_experience_json=json.dumps({}),
                notes=notes,
            )
            db.add(entry); db.commit(); db.refresh(entry)
            baseline = _get_or_init_baseline(patient_id, db)
            _update_baseline_from_entry(baseline, symptoms, db)
            imported += 1

    return {"imported": imported, "skipped": len(errors), "errors": errors[:20]}


# ─── PDF Clinical Notes Upload ─────────────────────────────────────────────────

@app.post("/upload-symptom-pdf/{patient_id}", tags=["Entries"])
async def upload_symptom_pdf(
    patient_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Extract symptom severities from a clinical notes PDF using NLP.
    Returns extracted symptoms with confidence scores before committing.
    Automatically creates a SymptomEntry from the extracted data.
    """
    from backend.nlp_extractor import extract_symptoms_from_text, extract_text_from_pdf

    patient = _get_patient_or_404(patient_id, db)
    config = DISEASE_CONFIGS.get(patient.disease, {})
    disease_symptoms = config.get("symptoms", [])

    pdf_bytes = await file.read()
    if not pdf_bytes:
        raise HTTPException(400, "Uploaded file is empty")

    try:
        text = extract_text_from_pdf(pdf_bytes)
    except RuntimeError as e:
        raise HTTPException(422, str(e))

    if not text.strip():
        raise HTTPException(422, "Could not extract any text from the PDF. Ensure it is not a scanned image-only PDF.")

    extracted = extract_symptoms_from_text(text, disease_symptoms, use_gemini=True)

    if not extracted:
        return {
            "extracted_symptoms": {},
            "confidence": {},
            "entry_id": None,
            "raw_text_preview": text[:500],
            "message": "No recognisable symptom mentions found in the PDF."
        }

    symptoms_dict = {sym: val for sym, (val, _conf) in extracted.items()}
    confidence_dict = {sym: conf for sym, (_val, conf) in extracted.items()}

    entry = SymptomEntry(
        patient_id=patient_id,
        timestamp=datetime.utcnow(),
        symptoms_json=json.dumps(symptoms_dict),
        triggers_json=json.dumps([]),
        lifestyle_json=json.dumps({}),
        shared_experience_json=json.dumps({}),
        notes=f"Extracted from PDF: {file.filename}"
    )
    db.add(entry); db.commit(); db.refresh(entry)
    baseline = _get_or_init_baseline(patient_id, db)
    _update_baseline_from_entry(baseline, symptoms_dict, db)

    return {
        "extracted_symptoms": symptoms_dict,
        "confidence": confidence_dict,
        "entry_id": entry.id,
        "raw_text_preview": text[:500],
        "message": f"Successfully extracted {len(symptoms_dict)} symptoms and saved entry."
    }


@app.get("/patients/{patient_id}/entries", response_model=List[SymptomEntryResponse], tags=["Entries"])
def get_entries(patient_id: str, limit: int = 100, db: Session = Depends(get_db)):
    _get_patient_or_404(patient_id, db)
    entries = (db.query(SymptomEntry)
               .filter(SymptomEntry.patient_id == patient_id)
               .order_by(SymptomEntry.timestamp.desc())
               .limit(limit).all())
    return [SymptomEntryResponse(
        id=e.id, patient_id=e.patient_id, timestamp=e.timestamp,
        symptoms=e.symptoms, triggers=e.triggers, lifestyle_context=e.lifestyle_context, shared_experience=e.shared_experience, notes=e.notes
    ) for e in entries]


# ─── Signal Computation ────────────────────────────────────────────────────────

@app.post("/compute-signals", response_model=SignalResponse, tags=["Signals"])
def compute_signals(body: SignalRequest, db: Session = Depends(get_db)):
    patient = _get_patient_or_404(body.patient_id, db)
    baseline = _get_or_init_baseline(body.patient_id, db)

    entries = (db.query(SymptomEntry)
               .filter(SymptomEntry.patient_id == body.patient_id)
               .order_by(SymptomEntry.timestamp.asc())
               .all())

    entry_dicts = _entries_to_dicts(entries)

    # Bootstrap baseline if needed
    if baseline.n_observations < 5 and entry_dicts:
        for sym in DISEASE_CONFIGS[patient.disease].get("symptoms", []):
            values = [e["symptoms"].get(sym) for e in entry_dicts if sym in e["symptoms"]]
            if values:
                mu, sig = initialize_baseline_from_history(values)
                mu_dict = baseline.mu
                sig_dict = baseline.sigma
                mu_dict[sym] = round(mu, 4)
                sig_dict[sym] = round(sig, 4)
                baseline.mu_json = json.dumps(mu_dict)
                baseline.sigma_json = json.dumps(sig_dict)
        db.commit()

    signals = compute_all_signals(
        entry_dicts,
        baseline.mu,
        baseline.sigma,
        patient.disease,
        body.window_days
    )

    # Persist computed signals
    cs = ComputedSignal(
        patient_id=body.patient_id,
        computed_at=datetime.utcnow(),
        z_scores_json=json.dumps({k: v.get("z_score", 0) for k, v in signals["z_scores"].items()}),
        volatility_index=signals["volatility"].get("value", 0.0),
        flare_acceleration=signals["flare_acceleration"].get("slope_7d", 0.0),
        trigger_correlations_json=json.dumps(signals["trigger_correlations"]),
        data_completeness=signals["missingness"].get("completeness_pct", 0) / 100.0,
        risk_category=signals["risk_category"],
        signals_suppressed_json=json.dumps(signals["missingness"].get("suppressed_signals", []))
    )
    db.add(cs)

    # Persist FIS
    fis = signals["functional_impact"]
    fs = FunctionalScore(
        patient_id=body.patient_id,
        computed_at=datetime.utcnow(),
        fis_composite=fis.get("composite", 0.0),
        domain_scores_json=json.dumps({
            k: fis.get(k, 0.0) for k in ["mobility", "cognitive", "sleep", "work", "social"]
        })
    )
    db.add(fs)

    # Create alert if needed
    if signals["risk_category"] in ("HIGH", "CRITICAL"):
        alert = Alert(
            patient_id=body.patient_id,
            alert_type=signals["risk_category"],
            risk_category=signals["risk_category"],
            message=f"Signal alert: {signals['risk_category']} risk detected. Red flags: {signals['red_flags']}"
        )
        db.add(alert)

    db.commit()

    # Build response
    from backend.schemas import (
        ZScoreDetail, VolatilityDetail, FlareAcceleration,
        TriggerCorrelation, FunctionalImpact, MissingnessDetail
    )

    z_resp = {k: ZScoreDetail(**v) for k, v in signals["z_scores"].items()}
    vol = signals["volatility"]
    vol_resp = VolatilityDetail(
        value=vol.get("value", 0.0),
        rolling_std=vol.get("rolling_std", 0.0),
        coefficient_of_variation=vol.get("cv", 0.0),
        approximate_entropy=vol.get("apen", 0.0),
        label=vol.get("label", "low")
    )
    fa = signals["flare_acceleration"]
    fa_resp = FlareAcceleration(
        slope_7d=fa.get("slope_7d", 0.0),
        slope_r2=fa.get("slope_r2", 0.0),
        change_point_detected=fa.get("change_point_detected", False),
        change_point_day=fa.get("change_point_day"),
        acceleration_label=fa.get("acceleration_label", "stable")
    )
    tc_resp = [TriggerCorrelation(**t) for t in signals["trigger_correlations"]]
    fi = signals["functional_impact"]
    fi_resp = FunctionalImpact(
        mobility=fi.get("mobility", 0.0),
        cognitive=fi.get("cognitive", 0.0),
        sleep=fi.get("sleep", 0.0),
        work=fi.get("work", 0.0),
        social=fi.get("social", 0.0),
        composite=fi.get("composite", 0.0),
        severity_label=fi.get("severity_label", "minimal")
    )
    miss = signals["missingness"]
    miss_resp = MissingnessDetail(
        completeness_pct=miss.get("completeness_pct", 0.0),
        entries_found=miss.get("entries_found", 0),
        entries_expected=miss.get("entries_expected", body.window_days),
        confidence_level=miss.get("confidence_level", "INSUFFICIENT"),
        suppressed_signals=miss.get("suppressed_signals", [])
    )

    return SignalResponse(
        patient_id=body.patient_id,
        computed_at=datetime.utcnow(),
        window_days=body.window_days,
        z_scores=z_resp,
        volatility=vol_resp,
        flare_acceleration=fa_resp,
        trigger_correlations=tc_resp,
        functional_impact=fi_resp,
        missingness=miss_resp,
        risk_category=signals["risk_category"],
        red_flags=signals["red_flags"]
    )


# ─── Risk Prediction ───────────────────────────────────────────────────────────

@app.post("/predict-risk", response_model=RiskPredictionResponse, tags=["ML"])
def predict_risk(body: RiskPredictionRequest, db: Session = Depends(get_db)):
    """
    Rule-based + signal-driven risk prediction.
    Returns risk probabilities derived from signal intensities.
    (Full ML model requires training data; this is the signal-based fallback for hackathon.)
    """
    patient = _get_patient_or_404(body.patient_id, db)
    latest_signal = (db.query(ComputedSignal)
                     .filter(ComputedSignal.patient_id == body.patient_id)
                     .order_by(ComputedSignal.computed_at.desc())
                     .first())
    latest_fis = (db.query(FunctionalScore)
                  .filter(FunctionalScore.patient_id == body.patient_id)
                  .order_by(FunctionalScore.computed_at.desc())
                  .first())
    latest_entry = (db.query(SymptomEntry)
                    .filter(SymptomEntry.patient_id == body.patient_id)
                    .order_by(SymptomEntry.timestamp.desc())
                    .first())

    if not latest_signal or not latest_entry:
        raise HTTPException(404, "No signals computed yet. Call /compute-signals first.")

    risk_payload = _compute_disease_aware_risk(patient, latest_entry, latest_signal, latest_fis)
    risk_cat = risk_payload["risk_category"]
    probs = risk_payload["risk_probabilities"]
    top_features = risk_payload["top_contributing_features"]

    # Simple 3-day forecast: assume trend continuation with mean reversion
    entries = (db.query(SymptomEntry)
               .filter(SymptomEntry.patient_id == body.patient_id)
               .order_by(SymptomEntry.timestamp.desc())
               .limit(7).all())
    relevant_symptoms = [
        s for s in DISEASE_CONFIGS[patient.disease].get("symptoms", [])
        if s not in {"sleep_quality", "stress_symptom_severity"}
    ][:5] + ["sleep_quality", "stress_symptom_severity"]

    if entries:
        recent_means = []
        for e in entries:
            vals = [e.symptoms.get(sym) for sym in relevant_symptoms if sym in e.symptoms]
            if vals:
                recent_means.append(sum(vals) / len(vals))
        recent_means.reverse()
        if recent_means:
            last = recent_means[-1]
            # Simple mean-reverting forecast
            target = sum(recent_means) / len(recent_means)
            forecast = [
                round(last + (target - last) * (i + 1) / 4, 2)
                for i in range(3)
            ]
        else:
            forecast = [5.0, 5.0, 5.0]
    else:
        forecast = [5.0, 5.0, 5.0]

    return RiskPredictionResponse(
        patient_id=body.patient_id,
        risk_category=risk_cat,
        risk_probabilities=probs,
        top_contributing_features=top_features,
        forecast_3d=forecast,
        confidence=latest_signal.data_completeness and
                   ("HIGH" if latest_signal.data_completeness > 0.8 else "MODERATE")
                   or "LOW"
    )


# ─── Summary Generation ────────────────────────────────────────────────────────

@app.post("/generate-summary", response_model=ClinicalSummaryResponse, tags=["Summary"])
def generate_summary(body: SummaryRequest, db: Session = Depends(get_db)):
    patient = _get_patient_or_404(body.patient_id, db)
    config = DISEASE_CONFIGS[patient.disease]

    # Get latest signals
    latest_signal = (db.query(ComputedSignal)
                     .filter(ComputedSignal.patient_id == body.patient_id)
                     .order_by(ComputedSignal.computed_at.desc())
                     .first())
    latest_fis = (db.query(FunctionalScore)
                  .filter(FunctionalScore.patient_id == body.patient_id)
                  .order_by(FunctionalScore.computed_at.desc())
                  .first())
    latest_entry = (db.query(SymptomEntry)
                    .filter(SymptomEntry.patient_id == body.patient_id)
                    .order_by(SymptomEntry.timestamp.desc())
                    .first())

    if not latest_signal:
        raise HTTPException(404, "No signals computed yet. Call /compute-signals first.")

    # Reconstruct signal dict from DB
    signals = {
        "z_scores": {
            sym: {"z_score": z, "value": 0, "baseline_mu": 0, "baseline_sigma": 0,
                  "interpretation": "elevated" if abs(z) > 1 else "stable"}
            for sym, z in latest_signal.z_scores.items()
        },
        "volatility": {
            "value": latest_signal.volatility_index or 0.0,
            "rolling_std": 0.0, "cv": 0.0, "apen": 0.0,
            "label": "high" if (latest_signal.volatility_index or 0) > 0.6
                     else "moderate" if (latest_signal.volatility_index or 0) > 0.3 else "low"
        },
        "flare_acceleration": {
            "slope_7d": latest_signal.flare_acceleration or 0.0,
            "slope_r2": 0.0, "change_point_detected": False,
            "change_point_day": None, "acceleration_label": "stable"
        },
        "trigger_correlations": latest_signal.trigger_correlations,
        "functional_impact": {
            "composite": latest_fis.fis_composite if latest_fis else 0.0,
            "severity_label": "minimal",
            **(latest_fis.domain_scores if latest_fis else {})
        },
        "missingness": {
            "completeness_pct": (latest_signal.data_completeness or 0) * 100,
            "entries_found": 0, "entries_expected": body.window_days,
            "confidence_level": "MODERATE", "suppressed_signals": latest_signal.signals_suppressed
        },
        "risk_category": latest_signal.risk_category,
        "red_flags": []
    }

    if latest_entry:
        risk_payload = _compute_disease_aware_risk(patient, latest_entry, latest_signal, latest_fis)
        signals["risk_category"] = risk_payload["risk_category"]
        signals["context_statement"] = risk_payload["context_statement"]

    summary = generate_structured_summary(
        patient_id=body.patient_id,
        disease=patient.disease,
        disease_name=config["name"],
        signals=signals,
        window_days=body.window_days
    )

    return ClinicalSummaryResponse(**summary)


@app.post("/generate-ai-explainer", response_model=DetailedReportResponse, tags=["Summary"])
def generate_ai_explainer(body: DetailedReportRequest, db: Session = Depends(get_db)):
    # Authenticate via patient check
    _get_patient_or_404(body.patient_id, db)
    try:
        report_text = generate_detailed_ai_report(
            patient_id=body.patient_id,
            disease_name=body.disease_name,
            deviations=body.deviations
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return DetailedReportResponse(report_text=report_text)


# ─── Semantic HPO Matching ──────────────────────────────────────────────────────

@app.post("/match-hpo", response_model=HPOMatchResponse, tags=["HPO"])
def semantic_hpo_match(body: HPOMatchRequest):
    """
    Semantically match a symptom key or free text to the nearest HPO term(s)
    using TF-IDF cosine similarity.
    """
    raw_matches = match_symptom(body.symptom_text, body.disease_id, body.top_k)
    return HPOMatchResponse(
        query=body.symptom_text,
        disease_id=body.disease_id,
        matches=[HPOMatchResult(**m) for m in raw_matches]
    )


@app.get("/diseases/{disease_id}/hpo-cluster", tags=["HPO"])
def disease_hpo_cluster(disease_id: str):
    """
    Return the full HPO semantic cluster for a disease —
    each symptom mapped to its closest HPO term and semantic neighbors.
    """
    if disease_id not in DISEASE_CONFIGS:
        raise HTTPException(404, f"Disease '{disease_id}' not found")
    return {"disease_id": disease_id, "cluster": get_disease_hpo_cluster(disease_id)}


# ─── Flare / Seizure Prediction Alert ─────────────────────────────────────────

@app.get("/flare-alert/{patient_id}", response_model=FlareAlertResponse, tags=["Alerts"])
def get_flare_alert(patient_id: str, db: Session = Depends(get_db)):
    """
    Predict days-to-next-flare for a patient using their latest computed signal.
    Derives a probability-weighted estimate from:
      - most recent risk category
      - symptom volatility index
      - z-score magnitude of primary deviations
    Returns an alert level and a user-visible message.
    """
    patient = _get_patient_or_404(patient_id, db)

    # Fetch latest computed signal
    latest_signal = (
        db.query(ComputedSignal)
        .filter(ComputedSignal.patient_id == patient_id)
        .order_by(ComputedSignal.computed_at.desc())
        .first()
    )

    if not latest_signal:
        return FlareAlertResponse(
            patient_id=patient_id,
            days_to_flare=30.0,
            alert_level="NORMAL",
            alert_color="#3fb950",
            message="No signals computed yet. Log symptoms to enable flare prediction.",
            confidence="LOW",
            based_on_risk="INSUFFICIENT_DATA"
        )

    risk = latest_signal.risk_category or "INSUFFICIENT_DATA"
    volatility = latest_signal.volatility_index or 0.0
    z_max = latest_signal.z_score_max or 0.0

    # Derive days-to-flare heuristic:
    # Base window from risk, refined by volatility and z-score magnitude
    base = {"CRITICAL": 2.0, "HIGH": 7.0, "MODERATE": 14.0, "LOW": 25.0, "INSUFFICIENT_DATA": 30.0}
    days = base.get(risk, 30.0)

    # Volatility shifts prediction earlier (high volatility = sooner flare)
    vol_adjustment = min(volatility * 5.0, 10.0)  # cap at -10 days
    z_adjustment = min(max(z_max - 1.0, 0.0) * 2.0, 8.0)  # z>1 → earlier
    days = max(1.0, days - vol_adjustment - z_adjustment)

    # Confidence: measured by how much data we have
    entry_count = db.query(SymptomEntry).filter(SymptomEntry.patient_id == patient_id).count()
    confidence = "HIGH" if entry_count >= 14 else "MODERATE" if entry_count >= 7 else "LOW"

    # Alert level thresholds
    if days <= 5:
        level, color = "CRITICAL", "#f85149"
        msg = f"Flare predicted in ~{days:.0f} day(s). Recommend urgent symptom review and care team contact."
    elif days <= 10:
        level, color = "WARNING", "#d29922"
        msg = f"Elevated flare risk — estimated ~{days:.0f} days. Consider proactive symptom logging."
    elif days <= 20:
        level, color = "WATCH", "#e3b341"
        msg = f"Flare may occur in ~{days:.0f} days. Maintain daily symptom tracking."
    else:
        level, color = "NORMAL", "#3fb950"
        msg = f"Low flare risk. Next predicted flare in ~{days:.0f} days. Keep up routine monitoring."

    return FlareAlertResponse(
        patient_id=patient_id,
        days_to_flare=round(days, 1),
        alert_level=level,
        alert_color=color,
        message=msg,
        confidence=confidence,
        based_on_risk=risk
    )


# ─── History ────────────────────────────────────────────────────────────────────

@app.get("/patients/{patient_id}/history", response_model=PatientHistoryResponse, tags=["History"])
def get_patient_history(patient_id: str, db: Session = Depends(get_db)):
    patient = _get_patient_or_404(patient_id, db)

    signals = (db.query(ComputedSignal)
               .filter(ComputedSignal.patient_id == patient_id)
               .order_by(ComputedSignal.computed_at.asc())
               .all())

    fis_scores = (db.query(FunctionalScore)
                  .filter(FunctionalScore.patient_id == patient_id)
                  .order_by(FunctionalScore.computed_at.asc())
                  .all())

    fis_map = {fs.computed_at.date().isoformat(): fs.fis_composite for fs in fis_scores}

    history = []
    for s in signals:
        date_str = s.computed_at.date().isoformat()
        z_max = max(s.z_scores.values()) if s.z_scores else 0.0
        history.append(SignalHistoryPoint(
            date=date_str,
            z_score_max=round(z_max, 4),
            volatility_index=round(s.volatility_index or 0.0, 4),
            fis_composite=round(fis_map.get(date_str, 0.0) or 0.0, 4),
            risk_category=s.risk_category,
            data_completeness=round((s.data_completeness or 0.0) * 100, 1)
        ))

    total_entries = db.query(SymptomEntry).filter(SymptomEntry.patient_id == patient_id).count()
    days_tracked = 0
    if signals:
        first = signals[0].computed_at
        last = signals[-1].computed_at
        days_tracked = (last - first).days + 1

    return PatientHistoryResponse(
        patient_id=patient_id,
        disease=patient.disease,
        history=history,
        total_entries=total_entries,
        days_tracked=days_tracked
    )


@app.get("/diseases/{disease_id}/shared-experiences-summary", response_model=SharedExperienceSummaryResponse, tags=["Community"])
def get_shared_experiences_summary(disease_id: str, db: Session = Depends(get_db)):
    if disease_id not in DISEASE_CONFIGS:
        raise HTTPException(404, f"Disease '{disease_id}' not found")

    recent_cutoff = datetime.utcnow() - timedelta(days=90)
    entries = (
        db.query(SymptomEntry)
        .join(Patient, SymptomEntry.patient_id == Patient.id)
        .filter(Patient.disease == disease_id)
        .filter(SymptomEntry.timestamp >= recent_cutoff)
        .order_by(SymptomEntry.timestamp.desc())
        .all()
    )

    helpful_counts = {}
    harder_counts = {}
    wisdom_snippets = []
    based_on_entries = 0

    for entry in entries:
        shared = entry.shared_experience or {}
        helped = shared.get("helpful_today") or []
        harder = shared.get("made_harder_today") or []
        snippet = _sanitize_shared_snippet(shared.get("wish_known_earlier"))

        if not helped and not harder and not snippet:
            continue

        based_on_entries += 1
        for item in helped:
            helpful_counts[item] = helpful_counts.get(item, 0) + 1
        for item in harder:
            harder_counts[item] = harder_counts.get(item, 0) + 1
        if snippet and snippet not in wisdom_snippets and len(wisdom_snippets) < 5:
            wisdom_snippets.append(snippet)

    helpful_counts = dict(sorted(helpful_counts.items(), key=lambda kv: kv[1], reverse=True)[:8])
    harder_counts = dict(sorted(harder_counts.items(), key=lambda kv: kv[1], reverse=True)[:8])

    return SharedExperienceSummaryResponse(
        disease_id=disease_id,
        disease_name=DISEASE_CONFIGS[disease_id]["name"],
        based_on_entries=based_on_entries,
        helpful_counts=helpful_counts,
        harder_counts=harder_counts,
        wisdom_snippets=wisdom_snippets,
    )


# ─── Disease Config ─────────────────────────────────────────────────────────────

@app.get("/diseases", tags=["Config"])
def list_diseases():
    return [
        {"id": k, "name": v["name"], "symptoms": v["symptoms"],
         "triggers": v["triggers"]}
        for k, v in DISEASE_CONFIGS.items()
    ]


@app.get("/diseases/{disease_id}", tags=["Config"])
def get_disease_config(disease_id: str):
    if disease_id not in DISEASE_CONFIGS:
        raise HTTPException(404, f"Disease '{disease_id}' not found")
    cfg = DISEASE_CONFIGS[disease_id]
    return {
        "id": disease_id,
        "name": cfg["name"],
        "symptoms": cfg["symptoms"],
        "symptom_labels": cfg.get("symptom_labels", {}),
        "triggers": cfg["triggers"],
        "fis_domain_weights": cfg.get("fis_domain_weights", {}),
        "domain_weights": cfg.get("domain_weights", {})
    }


# ─── Health Report ──────────────────────────────────────────────────────────────

@app.get("/health-report/{patient_id}", tags=["Reports"])
def get_health_report(patient_id: str, window_days: int = 7, lang: str = "en", db: Session = Depends(get_db)):
    """
    Generate a patient-friendly health report with:
    - Disease-specific recommended tests (evidence-based, 22 diseases)
    - Gemini AI summary of recent signal data (falls back gracefully)
    - Functional impact and risk overview
    """
    from backend.summary_generator import generate_health_report

    patient = _get_patient_or_404(patient_id, db)
    config = DISEASE_CONFIGS.get(patient.disease, {})
    disease_name = config.get("name", patient.disease)

    # Gather latest signals
    latest_signal = (
        db.query(ComputedSignal)
        .filter(ComputedSignal.patient_id == patient_id)
        .order_by(ComputedSignal.computed_at.desc())
        .first()
    )
    latest_fis = (
        db.query(FunctionalScore)
        .filter(FunctionalScore.patient_id == patient_id)
        .order_by(FunctionalScore.computed_at.desc())
        .first()
    )
    latest_entry = (
        db.query(SymptomEntry)
        .filter(SymptomEntry.patient_id == patient_id)
        .order_by(SymptomEntry.timestamp.desc())
        .first()
    )

    # Build signals dict
    signals_dict: dict = {}
    risk_data: dict = {}

    if latest_signal:
        z_scores_raw = latest_signal.z_scores or {}
        # Expand flat z-score dict to detail objects for generate_health_report
        signals_dict = {
            "risk_category": latest_signal.risk_category or "INSUFFICIENT_DATA",
            "z_scores": {
                k: {"z_score": v, "interpretation": (
                    "critical" if abs(v) >= 2.5 else
                    "elevated" if abs(v) >= 1.5 else
                    "watch" if abs(v) >= 1.0 else "stable"
                )}
                for k, v in z_scores_raw.items()
            },
            "volatility": {
                "value": latest_signal.volatility_index or 0.0,
                "label": (
                    "high" if (latest_signal.volatility_index or 0) > 0.7 else
                    "moderate" if (latest_signal.volatility_index or 0) > 0.35 else "low"
                ),
            },
            "flare_acceleration": {
                "slope_7d": latest_signal.flare_acceleration or 0.0,
                "acceleration_label": "stable",
            },
            "trigger_correlations": [],
            "functional_impact": {
                "composite": (latest_fis.fis_composite if latest_fis else 0.0),
                "mobility": (latest_fis.domain_scores.get("mobility", 0) if latest_fis else 0.0),
                "sleep": (latest_fis.domain_scores.get("sleep", 0) if latest_fis else 0.0),
                "cognitive": (latest_fis.domain_scores.get("cognitive", 0) if latest_fis else 0.0),
                "severity_label": "minimal",
            },
            "missingness": {"completeness_pct": 100, "entries_found": 1, "entries_expected": window_days},
            "red_flags": [],
        }

        if latest_entry and latest_signal:
            risk_payload = _compute_disease_aware_risk(patient, latest_entry, latest_signal, latest_fis)
            risk_data = {
                "risk_probabilities": risk_payload.get("risk_probabilities", {}),
                "forecast_3d": [],
            }
    else:
        signals_dict = {
            "risk_category": "INSUFFICIENT_DATA",
            "z_scores": {},
            "volatility": {"value": 0, "label": "low"},
            "flare_acceleration": {"slope_7d": 0, "acceleration_label": "stable"},
            "trigger_correlations": [],
            "functional_impact": {"composite": 0, "mobility": 0, "sleep": 0, "cognitive": 0, "severity_label": "minimal"},
            "missingness": {"completeness_pct": 0, "entries_found": 0, "entries_expected": window_days},
            "red_flags": [],
        }

    report = generate_health_report(
        patient_id=patient_id,
        disease=patient.disease,
        disease_name=disease_name,
        signals=signals_dict,
        risk_data=risk_data,
        window_days=window_days,
        lang=lang,
    )
    return report


# ─── Health ─────────────────────────────────────────────────────────────────────

@app.get("/health", tags=["Health"])
def health():
    return {"status": "ok", "version": "1.0.0", "timestamp": datetime.utcnow().isoformat()}



# ─── Baseline ───────────────────────────────────────────────────────────────────

@app.get("/patients/{patient_id}/baseline", tags=["Baseline"])
def get_baseline(patient_id: str, db: Session = Depends(get_db)):
    _get_patient_or_404(patient_id, db)
    baseline = _get_or_init_baseline(patient_id, db)
    return {
        "patient_id": patient_id,
        "mu": baseline.mu,
        "sigma": baseline.sigma,
        "n_observations": baseline.n_observations,
        "last_updated": baseline.last_updated,
        "alpha": baseline.alpha
    }
